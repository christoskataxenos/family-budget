from io import BytesIO
import os
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

# Εισαγωγή των μοντέλων και των εξαρτήσεων της εφαρμογής
from app.database import get_db
from app.models import Transaction, User
from app.security import get_current_user

# Εισαγωγή βιβλιοθηκών για παραγωγή XLSX
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Εισαγωγή βιβλιοθηκών για παραγωγή PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# Δημιουργία δρομολογητή για τις λειτουργίες εξαγωγής αρχείων
router = APIRouter(prefix="/api/v1/export", tags=["Export"])

# Καταχώρηση ελληνικής γραμματοσειράς για το ReportLab (χρήση Arial από τα Windows)
FONT_NAME = "Helvetica"
FONT_BOLD_NAME = "Helvetica-Bold"

# Έλεγχος αν υπάρχουν τα αρχεία γραμματοσειράς Arial στο σύστημα Windows
font_path_regular = "C:/Windows/Fonts/arial.ttf"
font_path_bold = "C:/Windows/Fonts/arialbd.ttf"

if os.path.exists(font_path_regular) and os.path.exists(font_path_bold):
    try:
        pdfmetrics.registerFont(TTFont("ArialGreek", font_path_regular))
        pdfmetrics.registerFont(TTFont("ArialGreek-Bold", font_path_bold))
        FONT_NAME = "ArialGreek"
        FONT_BOLD_NAME = "ArialGreek-Bold"
    except Exception:
        # Αν αποτύχει η εγγραφή της γραμματοσειράς, γίνεται fallback στη Helvetica
        pass


async def get_filtered_transactions(
    db: AsyncSession,
    current_user: User,
    transaction_type: Optional[str] = None,
    category: Optional[str] = None
) -> List[Transaction]:
    """
    Βοηθητική συνάρτηση ανάκτησης συναλλαγών βάσει δικαιωμάτων πρόσβασης και φίλτρων.
    """
    # Αρχικό ερώτημα ανάλογα με το ρόλο του χρήστη
    if current_user.role == "admin":
        query = select(Transaction)
    else:
        query = select(Transaction).where(
            (Transaction.user_id == current_user.id) | (Transaction.is_shared == True)
        )

    # Εφαρμογή επιπλέον φίλτρων αν έχουν οριστεί
    if transaction_type:
        query = query.where(Transaction.transaction_type == transaction_type)
    if category:
        query = query.where(Transaction.category == category)

    query = query.order_by(Transaction.date.desc())
    result = await db.execute(query)
    return result.scalars().all()


@router.get("/excel")
async def export_to_excel(
    transaction_type: Optional[str] = Query(None, description="Φίλτρο τύπου συναλλαγής (income/expense)"),
    category: Optional[str] = Query(None, description="Φίλτρο κατηγορίας"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Response:
    """
    Native εξαγωγή των συναλλαγών σε μορφή αρχείου Excel (.xlsx).
    """
    # Ανάκτηση των συναλλαγών από τη βάση δεδομένων
    transactions = await get_filtered_transactions(db, current_user, transaction_type, category)

    # Δημιουργία νέου βιβλίου εργασίας Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Συναλλαγές"

    # Ενεργοποίηση προβολής πλέγματος (grid lines)
    worksheet.views.sheetView[0].showGridLines = True

    # Ορισμός στυλ για την επικεφαλίδα (Header Styling)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ορισμός στυλ για τα δεδομένα
    data_font = Font(name="Calibri", size=11)
    number_font = Font(name="Calibri", size=11, bold=False)
    thin_border_side = Side(border_style="thin", color="E5E7EB")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Λίστα επικεφαλίδων
    headers = [
        "ID", "Ημερομηνία", "Τύπος", "Κατηγορία", 
        "Ποσό (€)", "Περιγραφή", "Συχνότητα", "Κοινόχρηστη"
    ]
    worksheet.append(headers)

    # Μορφοποίηση γραμμής επικεφαλίδων
    for col_num in range(1, len(headers) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = cell_border

    # Αρχικοποίηση αθροιστών
    total_income = 0.0
    total_expense = 0.0

    # Εγγραφή των συναλλαγών στις γραμμές του Excel
    row_idx = 2
    for item in transactions:
        type_str = "Έσοδο" if item.transaction_type == "income" else "Έξοδο"
        is_shared_str = "Ναι" if item.is_shared else "Όχι"
        formatted_date = item.date.strftime("%Y-%m-%d") if item.date else ""
        amount_val = float(item.amount)

        if item.transaction_type == "income":
            total_income += amount_val
        else:
            total_expense += amount_val

        row_data = [
            item.id,
            formatted_date,
            type_str,
            item.category,
            amount_val,
            item.description or "",
            item.frequency,
            is_shared_str
        ]
        worksheet.append(row_data)

        # Μορφοποίηση κελιών της γραμμής
        for col_idx in range(1, len(row_data) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = cell_border

            # Μορφοποίηση στήλης Ποσού (Currency €)
            if col_idx == 5:
                cell.number_format = '€#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                cell.font = number_font
            elif col_idx in [1, 2, 3, 7, 8]:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

        row_idx += 1

    # Προσθήκη γραμμής συνόλων στο τέλος
    row_idx += 1
    worksheet.cell(row=row_idx, column=1, value="Σύνολο Εσόδων:").font = Font(name="Calibri", size=11, bold=True)
    income_cell = worksheet.cell(row=row_idx, column=5, value=total_income)
    income_cell.font = Font(name="Calibri", size=11, bold=True, color="16A34A")
    income_cell.number_format = '€#,##0.00'

    row_idx += 1
    worksheet.cell(row=row_idx, column=1, value="Σύνολο Εξόδων:").font = Font(name="Calibri", size=11, bold=True)
    expense_cell = worksheet.cell(row=row_idx, column=5, value=total_expense)
    expense_cell.font = Font(name="Calibri", size=11, bold=True, color="DC2626")
    expense_cell.number_format = '€#,##0.00'

    row_idx += 1
    worksheet.cell(row=row_idx, column=1, value="Καθαρό Υπόλοιπο:").font = Font(name="Calibri", size=11, bold=True)
    net_cell = worksheet.cell(row=row_idx, column=5, value=total_income - total_expense)
    net_cell.font = Font(name="Calibri", size=11, bold=True)
    net_cell.number_format = '€#,##0.00'

    # Αυτόματη ρύθμιση πλάτους στηλών (Auto Column Width)
    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Αποθήκευση του αρχείου σε μνήμη (BytesIO buffer)
    output_stream = BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)

    # Επιστροφή απάντησης με το binary αρχείο Excel
    headers_response = {
        "Content-Disposition": 'attachment; filename="family_budget_transactions.xlsx"'
    }
    return Response(
        content=output_stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_response
    )


@router.get("/pdf")
async def export_to_pdf(
    transaction_type: Optional[str] = Query(None, description="Φίλτρο τύπου συναλλαγής (income/expense)"),
    category: Optional[str] = Query(None, description="Φίλτρο κατηγορίας"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Response:
    """
    Native εξαγωγή των συναλλαγών σε μορφή αρχείου PDF.
    """
    # Ανάκτηση των συναλλαγών από τη βάση δεδομένων
    transactions = await get_filtered_transactions(db, current_user, transaction_type, category)

    # Δημιουργία μνήμης buffer για το αρχείο PDF
    buffer = BytesIO()

    # Ρύθμιση εγγράφου PDF σε οριζόντια διάταξη A4 (Landscape)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    story = []
    styles = getSampleStyleSheet()

    # Ορισμός εξατομικευμένων στυλ κειμένου
    title_style = ParagraphStyle(
        name="ReportTitle",
        fontName=FONT_BOLD_NAME,
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#1E3A8A"),
        alignment=0,
        spaceAfter=6
    )

    subtitle_style = ParagraphStyle(
        name="ReportSubtitle",
        fontName=FONT_NAME,
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#6B7280"),
        spaceAfter=15
    )

    cell_style_head = ParagraphStyle(
        name="CellHead",
        fontName=FONT_BOLD_NAME,
        fontSize=10,
        leading=12,
        textColor=colors.white,
        alignment=1
    )

    cell_style_body = ParagraphStyle(
        name="CellBody",
        fontName=FONT_NAME,
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#1F2937")
    )

    cell_style_body_center = ParagraphStyle(
        name="CellBodyCenter",
        fontName=FONT_NAME,
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#1F2937"),
        alignment=1
    )

    cell_style_body_right = ParagraphStyle(
        name="CellBodyRight",
        fontName=FONT_NAME,
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#1F2937"),
        alignment=2
    )

    # Προσθήκη τίτλου και πληροφοριών αναφοράς
    story.append(Paragraph("Family Budget - Αναφορά Συναλλαγών", title_style))
    story.append(Paragraph(f"Χρήστης: {current_user.full_name} ({current_user.email}) | Ρόλος: {current_user.role}", subtitle_style))
    story.append(Spacer(1, 10))

    # Πίνακας δεδομένων
    table_data = [
        [
            Paragraph("ID", cell_style_head),
            Paragraph("Ημερομηνία", cell_style_head),
            Paragraph("Τύπος", cell_style_head),
            Paragraph("Κατηγορία", cell_style_head),
            Paragraph("Ποσό (€)", cell_style_head),
            Paragraph("Περιγραφή", cell_style_head),
            Paragraph("Συχνότητα", cell_style_head),
            Paragraph("Κοινόχρηστη", cell_style_head)
        ]
    ]

    total_income = 0.0
    total_expense = 0.0

    for item in transactions:
        type_str = "Έσοδο" if item.transaction_type == "income" else "Έξοδο"
        is_shared_str = "Ναι" if item.is_shared else "Όχι"
        formatted_date = item.date.strftime("%Y-%m-%d") if item.date else ""
        amount_val = float(item.amount)

        if item.transaction_type == "income":
            total_income += amount_val
        else:
            total_expense += amount_val

        amount_formatted = f"€{amount_val:,.2f}"

        row = [
            Paragraph(str(item.id), cell_style_body_center),
            Paragraph(formatted_date, cell_style_body_center),
            Paragraph(type_str, cell_style_body_center),
            Paragraph(item.category or "", cell_style_body),
            Paragraph(amount_formatted, cell_style_body_right),
            Paragraph(item.description or "-", cell_style_body),
            Paragraph(item.frequency or "", cell_style_body_center),
            Paragraph(is_shared_str, cell_style_body_center)
        ]
        table_data.append(row)

    # Πλάτος στηλών για landscape A4 (συνολικό πλάτος ~780pt)
    col_widths = [40, 75, 60, 100, 80, 245, 90, 70]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Στυλ πίνακα (TableStyle)
    ts = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
    ]

    # Εναλλαγή χρωμάτων στις γραμμές (Zebra striping)
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            ts.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F9FAFB")))

    table.setStyle(TableStyle(ts))
    story.append(table)
    story.append(Spacer(1, 15))

    # Σύνολα στο τέλος
    summary_data = [
        [
            Paragraph("<b>Σύνολο Εσόδων:</b>", cell_style_body),
            Paragraph(f"<b>€{total_income:,.2f}</b>", ParagraphStyle(name="Inc", parent=cell_style_body_right, textColor=colors.HexColor("#16A34A")))
        ],
        [
            Paragraph("<b>Σύνολο Εξόδων:</b>", cell_style_body),
            Paragraph(f"<b>€{total_expense:,.2f}</b>", ParagraphStyle(name="Exp", parent=cell_style_body_right, textColor=colors.HexColor("#DC2626")))
        ],
        [
            Paragraph("<b>Καθαρό Υπόλοιπο:</b>", cell_style_body),
            Paragraph(f"<b>€{(total_income - total_expense):,.2f}</b>", ParagraphStyle(name="Net", parent=cell_style_body_right, textColor=colors.HexColor("#1E3A8A")))
        ]
    ]
    summary_table = Table(summary_data, colWidths=[120, 100])
    summary_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_table)

    # Κατασκευή PDF εγγράφου
    doc.build(story)

    # Επιστροφή PDF απάντησης
    pdf_bytes = buffer.getvalue()
    buffer.close()

    headers_response = {
        "Content-Disposition": 'attachment; filename="family_budget_transactions.pdf"'
    }
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers=headers_response
    )
